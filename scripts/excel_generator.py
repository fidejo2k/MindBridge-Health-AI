import anthropic
import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def analyze_patient_from_csv(patient_row, client):
    """Analyze a patient from CSV data"""
    
    patient_summary = f"""
Patient ID: {patient_row['patient_id']}
Name: {patient_row['name']}
Last Appointment: {patient_row['last_appointment']}
Appointments Missed (last 6 months): {patient_row['appointments_missed']}
Medication Adherence: {float(patient_row['medication_adherence'])*100:.0f}%
Crisis Calls (30 days): {patient_row['crisis_calls_30days']}
Diagnosis: {patient_row['diagnosis']}
Case Manager: {patient_row['case_manager']}
"""
    
    prompt = f"""
You are a clinical risk assessment assistant.

Analyze this patient and provide:
1. Risk Level (Low/Medium/High)
2. Primary Risk Factor (max 100 characters)
3. Recommended Action (max 150 characters)

PATIENT DATA:
{patient_summary}

Return in this exact format:
Risk Level: [level]
Primary Factor: [factor]
Action: [action]
"""
    
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}]
    )
    
    return message.content[0].text

def create_excel_report(csv_file, output_file):
    """Create a professional Excel spreadsheet report"""
    
    # Initialize Claude client
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
    
    # Create Excel workbook
    wb = Workbook()
    
    # Read patient data
    print(f"Reading patient data from: {csv_file}")
    patients = []
    
    with open(csv_file, 'r', encoding='utf-8') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            patients.append(row)
    
    print(f"✓ Loaded {len(patients)} patients\n")
    
    # Analyze all patients first
    patient_analyses = []
    high_risk_patients = []
    medium_risk_patients = []
    low_risk_patients = []
    
    print("Analyzing patients...\n")
    
    for i, patient in enumerate(patients, 1):
        print(f"[{i}/{len(patients)}] Analyzing {patient['name']}...")
        
        # Get AI analysis
        analysis = analyze_patient_from_csv(patient, client)
        
        # Parse analysis
        risk_level = "Unknown"
        primary_factor = "Not analyzed"
        action = "No action specified"
        
        for line in analysis.split('\n'):
            if line.startswith('Risk Level:'):
                risk_level = line.replace('Risk Level:', '').strip()
            elif line.startswith('Primary Factor:'):
                primary_factor = line.replace('Primary Factor:', '').strip()
            elif line.startswith('Action:'):
                action = line.replace('Action:', '').strip()
        
        # Store analysis
        patient_analysis = {
            'patient': patient,
            'risk_level': risk_level,
            'primary_factor': primary_factor,
            'action': action
        }
        patient_analyses.append(patient_analysis)
        
        # Categorize by risk
        if "High" in risk_level:
            high_risk_patients.append(patient_analysis)
        elif "Medium" in risk_level:
            medium_risk_patients.append(patient_analysis)
        else:
            low_risk_patients.append(patient_analysis)
    
    print(f"\n✓ Analysis complete!")
    print(f"  High Risk: {len(high_risk_patients)}")
    print(f"  Medium Risk: {len(medium_risk_patients)}")
    print(f"  Low Risk: {len(low_risk_patients)}\n")
    
    # Define color fills
    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    green_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Define fonts
    header_font = Font(bold=True, color="FFFFFF", size=12)
    bold_font = Font(bold=True)
    
    # Define alignment
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Helper function to format header row
    def format_header_row(sheet, row_num):
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_num, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
    
    # Helper function to add patient row
    def add_patient_row(sheet, row_num, patient_analysis):
        patient = patient_analysis['patient']
        risk_level = patient_analysis['risk_level']
        
        # Determine fill color
        if "High" in risk_level:
            fill = red_fill
        elif "Medium" in risk_level:
            fill = orange_fill
        else:
            fill = green_fill
        
        # Add data
        sheet.cell(row_num, 1, patient['patient_id'])
        sheet.cell(row_num, 2, patient['name'])
        sheet.cell(row_num, 3, risk_level)
        sheet.cell(row_num, 4, patient['diagnosis'])
        sheet.cell(row_num, 5, patient['case_manager'])
        sheet.cell(row_num, 6, patient['last_appointment'])
        sheet.cell(row_num, 7, int(patient['appointments_missed']))
        sheet.cell(row_num, 8, f"{float(patient['medication_adherence'])*100:.0f}%")
        sheet.cell(row_num, 9, int(patient['crisis_calls_30days']))
        sheet.cell(row_num, 10, patient_analysis['primary_factor'])
        sheet.cell(row_num, 11, patient_analysis['action'])
        
        # Apply formatting
        for col_num in range(1, 12):
            cell = sheet.cell(row_num, col_num)
            cell.fill = fill
            cell.border = thin_border
            if col_num in [3, 7, 8, 9]:  # Center align some columns
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align
    
    print("Creating Excel sheets...\n")
    
    # SHEET 1: SUMMARY
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = "OAKWOOD BEHAVIORAL HEALTH"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = "Daily Patient Risk Screening Report"
    ws_summary['A2'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A2:D2')
    
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%A, %B %d, %Y at %I:%M %p')}"
    ws_summary.merge_cells('A3:D3')
    
    # Statistics
    ws_summary['A5'] = "Risk Level"
    ws_summary['B5'] = "Count"
    ws_summary['C5'] = "Percentage"
    ws_summary['D5'] = "Status"
    format_header_row(ws_summary, 5)
    
    ws_summary['A6'] = "HIGH RISK"
    ws_summary['B6'] = len(high_risk_patients)
    ws_summary['C6'] = f"{len(high_risk_patients)/len(patients)*100:.1f}%"
    ws_summary['D6'] = "IMMEDIATE ATTENTION"
    ws_summary['A6'].fill = red_fill
    ws_summary['B6'].fill = red_fill
    ws_summary['C6'].fill = red_fill
    ws_summary['D6'].fill = red_fill
    ws_summary['A6'].font = bold_font
    
    ws_summary['A7'] = "MEDIUM RISK"
    ws_summary['B7'] = len(medium_risk_patients)
    ws_summary['C7'] = f"{len(medium_risk_patients)/len(patients)*100:.1f}%"
    ws_summary['D7'] = "FOLLOW-UP 48-72 HRS"
    ws_summary['A7'].fill = orange_fill
    ws_summary['B7'].fill = orange_fill
    ws_summary['C7'].fill = orange_fill
    ws_summary['D7'].fill = orange_fill
    ws_summary['A7'].font = bold_font
    
    ws_summary['A8'] = "LOW RISK"
    ws_summary['B8'] = len(low_risk_patients)
    ws_summary['C8'] = f"{len(low_risk_patients)/len(patients)*100:.1f}%"
    ws_summary['D8'] = "ROUTINE MONITORING"
    ws_summary['A8'].fill = green_fill
    ws_summary['B8'].fill = green_fill
    ws_summary['C8'].fill = green_fill
    ws_summary['D8'].fill = green_fill
    ws_summary['A8'].font = bold_font
    
    ws_summary['A9'] = "TOTAL"
    ws_summary['B9'] = len(patients)
    ws_summary['C9'] = "100%"
    ws_summary['D9'] = ""
    ws_summary['A9'].font = bold_font
    ws_summary['B9'].font = bold_font
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 25
    
    print("  ✓ Summary sheet created")
    
    # SHEET 2: ALL PATIENTS
    ws_all = wb.create_sheet("All Patients")
    
    # Headers
    headers = ["Patient ID", "Name", "Risk Level", "Diagnosis", "Case Manager", 
               "Last Appt", "Missed Appts", "Med Adherence", "Crisis Calls",
               "Primary Risk Factor", "Recommended Action"]
    
    for col_num, header in enumerate(headers, 1):
        ws_all.cell(1, col_num, header)
    
    format_header_row(ws_all, 1)
    
    # Add all patients
    for i, patient_analysis in enumerate(patient_analyses, 2):
        add_patient_row(ws_all, i, patient_analysis)
    
    # Adjust column widths
    ws_all.column_dimensions['A'].width = 12
    ws_all.column_dimensions['B'].width = 20
    ws_all.column_dimensions['C'].width = 15
    ws_all.column_dimensions['D'].width = 25
    ws_all.column_dimensions['E'].width = 18
    ws_all.column_dimensions['F'].width = 15
    ws_all.column_dimensions['G'].width = 13
    ws_all.column_dimensions['H'].width = 15
    ws_all.column_dimensions['I'].width = 13
    ws_all.column_dimensions['J'].width = 35
    ws_all.column_dimensions['K'].width = 45
    
    # Freeze top row
    ws_all.freeze_panes = 'A2'
    
    print("  ✓ All Patients sheet created")
    
    # SHEET 3: HIGH RISK ONLY
    if high_risk_patients:
        ws_high = wb.create_sheet("High Risk")
        
        for col_num, header in enumerate(headers, 1):
            ws_high.cell(1, col_num, header)
        
        format_header_row(ws_high, 1)
        
        for i, patient_analysis in enumerate(high_risk_patients, 2):
            add_patient_row(ws_high, i, patient_analysis)
        
        # Same column widths as All Patients
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws_high.column_dimensions[col].width = ws_all.column_dimensions[col].width
        
        ws_high.freeze_panes = 'A2'
        
        print("  ✓ High Risk sheet created")
    
    # SHEET 4: MEDIUM RISK ONLY
    if medium_risk_patients:
        ws_medium = wb.create_sheet("Medium Risk")
        
        for col_num, header in enumerate(headers, 1):
            ws_medium.cell(1, col_num, header)
        
        format_header_row(ws_medium, 1)
        
        for i, patient_analysis in enumerate(medium_risk_patients, 2):
            add_patient_row(ws_medium, i, patient_analysis)
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws_medium.column_dimensions[col].width = ws_all.column_dimensions[col].width
        
        ws_medium.freeze_panes = 'A2'
        
        print("  ✓ Medium Risk sheet created")
    
    # SHEET 5: LOW RISK ONLY
    if low_risk_patients:
        ws_low = wb.create_sheet("Low Risk")
        
        for col_num, header in enumerate(headers, 1):
            ws_low.cell(1, col_num, header)
        
        format_header_row(ws_low, 1)
        
        for i, patient_analysis in enumerate(low_risk_patients, 2):
            add_patient_row(ws_low, i, patient_analysis)
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws_low.column_dimensions[col].width = ws_all.column_dimensions[col].width
        
        ws_low.freeze_panes = 'A2'
        
        print("  ✓ Low Risk sheet created")
    
    # Save workbook
    wb.save(output_file)
    
    print(f"\n✓ EXCEL SPREADSHEET CREATED!")
    print(f"✓ File saved: {output_file}")
    print(f"\n📊 SHEETS CREATED:")
    print(f"   1. Summary (statistics)")
    print(f"   2. All Patients ({len(patients)} total)")
    print(f"   3. High Risk ({len(high_risk_patients)} patients)")
    print(f"   4. Medium Risk ({len(medium_risk_patients)} patients)")
    print(f"   5. Low Risk ({len(low_risk_patients)} patients)")
    
    return output_file

# Run the script
if __name__ == "__main__":
    input_csv = "patients.csv"
    output_xlsx = f"reports/patient_screening_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print("=" * 70)
    print("EXCEL SPREADSHEET GENERATOR")
    print("=" * 70)
    print()
    
    create_excel_report(input_csv, output_xlsx)
    
    print(f"\n📊 Open the Excel file: {output_xlsx}")
    print("\nYou can now SORT, FILTER, and ANALYZE this data!")
